"""
to_excel.py（COSCO NEUR 線）
────────────────────────────
從 COSCO NEUR PDF 萃取 Ocean Freight rate 並輸出成 Excel，方便檢查萃取結果。

用法：python to_excel.py <path_to_pdf>
輸出：pdf_extract.xlsx（執行目錄）

依賴：pip install pdfplumber openpyxl
"""

import sys
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

try:
    from cosco_parser import extract_ocean_rates, extract_us_inland
except ImportError:
    from cosco_neur.cosco_parser import extract_ocean_rates, extract_us_inland


HEADER_FILL = PatternFill("solid", start_color="1F3864")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
DATA_FONT   = Font(name="Arial", size=10)
CENTER      = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT        = Alignment(horizontal="left",   vertical="center", wrap_text=True)


def write_rate_sheet(ws, rows):
    ws.title = "Ocean Freight Rates"
    for col, h in enumerate(["POL", "POD", "20DV (USD)", "40DV/40HQ (USD)"], 1):
        cell = ws.cell(1, col, h)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = CENTER
    ws.row_dimensions[1].height = 20
    for r_idx, row in enumerate(rows, 2):
        for c_idx, key in enumerate(["pol", "pod", "rate_20", "rate_40"], 1):
            cell = ws.cell(r_idx, c_idx, row[key])
            cell.font      = DATA_FONT
            cell.alignment = LEFT if c_idx <= 2 else CENTER
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16


def write_inland_sheet(ws, rows):
    ws.title = "US Inland"
    headers = ["Location", "Location Type", "Routing Via", "20DV (USD)", "40DV/40HQ (USD)", "40RF/40RQ (USD)"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(1, col, h)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = CENTER
    ws.row_dimensions[1].height = 20
    keys = ["location", "location_type", "routing_via", "rate_20", "rate_40", "rate_40rf"]
    for r_idx, row in enumerate(rows, 2):
        for c_idx, key in enumerate(keys, 1):
            val = row.get(key)
            cell = ws.cell(r_idx, c_idx, val)
            cell.font      = DATA_FONT
            cell.alignment = LEFT if c_idx <= 3 else CENTER
    for col, width in zip("ABCDEF", [20, 18, 14, 14, 16, 16]):
        ws.column_dimensions[col].width = width


def main():
    if len(sys.argv) < 2:
        print("Usage: python to_excel.py <path_to_pdf>")
        sys.exit(1)

    pdf_path = sys.argv[1]
    rates = extract_ocean_rates(pdf_path)
    ramp_rows = extract_us_inland(pdf_path)

    wb = Workbook()
    write_rate_sheet(wb.active, rates)

    if ramp_rows:
        write_inland_sheet(wb.create_sheet(), ramp_rows)

    out_path = os.path.join(os.getcwd(), "pdf_extract.xlsx")
    wb.save(out_path)

    summary = f"Saved: {out_path} ({len(rates)} ocean freight rows"
    summary += f", {len(ramp_rows)} US inland rows)" if ramp_rows else ", no US inland table in this PDF)"
    print(summary)


if __name__ == "__main__":
    main()
